"""Utility functions for attendance reporting and data export."""

import csv
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

try:
	import openpyxl
	from openpyxl.styles import Font, PatternFill, Alignment
	OPENPYXL_AVAILABLE = True
except ImportError:
	OPENPYXL_AVAILABLE = False

try:
	from reportlab.lib import colors
	from reportlab.lib.pagesizes import letter, A4
	from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
	from reportlab.lib.units import inch
	from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
	REPORTLAB_AVAILABLE = True
except ImportError:
	REPORTLAB_AVAILABLE = False


class ReportGenerator:
	"""Generate custom attendance reports in various formats."""
	
	def __init__(self, queryset, fields, grouping=None, sorting=None, include_summary=True):
		"""
		Initialize the report generator.
		
		Args:
			queryset: AttendanceRecord queryset to report on
			fields: List of field names to include
			grouping: Grouping option (STUDENT, CLASSROOM, SUBJECT, DATE, NONE)
			sorting: Sorting option
			include_summary: Whether to include summary statistics
		"""
		self.queryset = queryset
		self.fields = fields
		self.grouping = grouping
		self.sorting = sorting
		self.include_summary = include_summary
	
	def get_field_value(self, record, field_name):
		"""Extract the value for a given field from a record."""
		field_mapping = {
			'date': lambda r: r.date.strftime('%Y-%m-%d'),
			'student_admission_number': lambda r: r.student.admission_number,
			'student_name': lambda r: r.student.full_name,
			'classroom': lambda r: str(r.classroom),
			'subject': lambda r: str(r.subject) if r.subject else 'N/A',
			'status': lambda r: r.get_status_display(),
			'notes': lambda r: r.notes or '',
			'recorded_by': lambda r: r.recorded_by.get_full_name() if r.recorded_by else 'System',
			'recorded_at': lambda r: timezone.localtime(r.recorded_at).strftime('%Y-%m-%d %H:%M'),
		}
		return field_mapping.get(field_name, lambda r: '')(record)
	
	def get_headers(self):
		"""Get human-readable headers for selected fields."""
		header_mapping = {
			'date': 'Date',
			'student_admission_number': 'Admission Number',
			'student_name': 'Student Name',
			'classroom': 'Classroom',
			'subject': 'Subject',
			'status': 'Status',
			'notes': 'Notes',
			'recorded_by': 'Recorded By',
			'recorded_at': 'Recorded At',
		}
		return [header_mapping.get(field, field) for field in self.fields]
	
	def apply_sorting(self, queryset):
		"""Apply sorting to the queryset."""
		sorting_mapping = {
			'DATE_ASC': 'date',
			'DATE_DESC': '-date',
			'STUDENT_ASC': 'student__first_name',
			'STUDENT_DESC': '-student__first_name',
			'STATUS_ASC': 'status',
			'STATUS_DESC': '-status',
		}
		if self.sorting and self.sorting in sorting_mapping:
			return queryset.order_by(sorting_mapping[self.sorting])
		return queryset
	
	def calculate_summary(self):
		"""Calculate summary statistics."""
		from .models import AttendanceRecord
		
		total = self.queryset.count()
		if total == 0:
			return None
		
		present_count = self.queryset.filter(status=AttendanceRecord.Status.PRESENT).count()
		absent_count = self.queryset.filter(status=AttendanceRecord.Status.ABSENT).count()
		late_count = self.queryset.filter(status=AttendanceRecord.Status.LATE).count()
		excused_count = self.queryset.filter(status=AttendanceRecord.Status.EXCUSED).count()
		
		return {
			'total': total,
			'present': present_count,
			'absent': absent_count,
			'late': late_count,
			'excused': excused_count,
			'present_pct': (present_count / total * 100) if total > 0 else 0,
			'absent_pct': (absent_count / total * 100) if total > 0 else 0,
		}
	
	def export_csv(self):
		"""Export report as CSV."""
		response = HttpResponse(content_type='text/csv')
		response['Content-Disposition'] = f'attachment; filename="attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
		
		writer = csv.writer(response)
		
		# Write headers
		writer.writerow(self.get_headers())
		
		# Apply sorting
		queryset = self.apply_sorting(self.queryset)
		
		# Write data rows
		for record in queryset:
			row = [self.get_field_value(record, field) for field in self.fields]
			writer.writerow(row)
		
		# Write summary if requested
		if self.include_summary:
			summary = self.calculate_summary()
			if summary:
				writer.writerow([])  # Empty row
				writer.writerow(['Summary Statistics'])
				writer.writerow(['Total Records', summary['total']])
				writer.writerow(['Present', f"{summary['present']} ({summary['present_pct']:.1f}%)"])
				writer.writerow(['Absent', f"{summary['absent']} ({summary['absent_pct']:.1f}%)"])
				writer.writerow(['Late', summary['late']])
				writer.writerow(['Excused', summary['excused']])
		
		return response
	
	def export_xlsx(self):
		"""Export report as Excel (XLSX)."""
		if not OPENPYXL_AVAILABLE:
			raise ImportError('openpyxl is required for Excel export. Install it with: pip install openpyxl')
		
		# Create workbook
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = 'Attendance Report'
		
		# Style definitions
		header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
		header_font = Font(color='FFFFFF', bold=True)
		header_alignment = Alignment(horizontal='center', vertical='center')
		
		# Write headers
		headers = self.get_headers()
		for col_num, header in enumerate(headers, 1):
			cell = ws.cell(row=1, column=col_num, value=header)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = header_alignment
		
		# Apply sorting
		queryset = self.apply_sorting(self.queryset)
		
		# Write data rows
		for row_num, record in enumerate(queryset, 2):
			for col_num, field in enumerate(self.fields, 1):
				value = self.get_field_value(record, field)
				ws.cell(row=row_num, column=col_num, value=value)
		
		# Auto-adjust column widths
		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = min(max_length + 2, 50)
			ws.column_dimensions[column_letter].width = adjusted_width
		
		# Write summary if requested
		if self.include_summary:
			summary = self.calculate_summary()
			if summary:
				row_num = ws.max_row + 2
				ws.cell(row=row_num, column=1, value='Summary Statistics').font = Font(bold=True)
				row_num += 1
				ws.cell(row=row_num, column=1, value='Total Records')
				ws.cell(row=row_num, column=2, value=summary['total'])
				row_num += 1
				ws.cell(row=row_num, column=1, value='Present')
				ws.cell(row=row_num, column=2, value=f"{summary['present']} ({summary['present_pct']:.1f}%)")
				row_num += 1
				ws.cell(row=row_num, column=1, value='Absent')
				ws.cell(row=row_num, column=2, value=f"{summary['absent']} ({summary['absent_pct']:.1f}%)")
				row_num += 1
				ws.cell(row=row_num, column=1, value='Late')
				ws.cell(row=row_num, column=2, value=summary['late'])
				row_num += 1
				ws.cell(row=row_num, column=1, value='Excused')
				ws.cell(row=row_num, column=2, value=summary['excused'])
		
		# Save to BytesIO
		output = BytesIO()
		wb.save(output)
		output.seek(0)
		
		response = HttpResponse(
			output.read(),
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
		response['Content-Disposition'] = f'attachment; filename="attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
		
		return response
	
	def export_pdf(self):
		"""Export report as PDF."""
		if not REPORTLAB_AVAILABLE:
			raise ImportError('reportlab is required for PDF export. Install it with: pip install reportlab')
		
		buffer = BytesIO()
		doc = SimpleDocTemplate(buffer, pagesize=letter)
		elements = []
		styles = getSampleStyleSheet()
		
		# Title
		title_style = ParagraphStyle(
			'CustomTitle',
			parent=styles['Heading1'],
			fontSize=18,
			textColor=colors.HexColor('#366092'),
			spaceAfter=30,
			alignment=1  # Center
		)
		elements.append(Paragraph('Attendance Report', title_style))
		elements.append(Spacer(1, 0.2 * inch))
		
		# Metadata
		meta_style = styles['Normal']
		elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', meta_style))
		elements.append(Paragraph(f'Total Records: {self.queryset.count()}', meta_style))
		elements.append(Spacer(1, 0.3 * inch))
		
		# Apply sorting
		queryset = self.apply_sorting(self.queryset)
		
		# Prepare table data
		headers = self.get_headers()
		table_data = [headers]
		
		for record in queryset[:100]:  # Limit to 100 records for PDF
			row = [self.get_field_value(record, field) for field in self.fields]
			table_data.append(row)
		
		# Create table
		table = Table(table_data)
		table.setStyle(TableStyle([
			('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
			('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
			('ALIGN', (0, 0), (-1, -1), 'LEFT'),
			('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, 0), 10),
			('BOTTOMPADDING', (0, 0), (-1, 0), 12),
			('BACKGROUND', (0, 1), (-1, -1), colors.beige),
			('GRID', (0, 0), (-1, -1), 1, colors.black),
			('FONTSIZE', (0, 1), (-1, -1), 8),
			('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
		]))
		elements.append(table)
		
		# Add summary if requested
		if self.include_summary:
			summary = self.calculate_summary()
			if summary:
				elements.append(Spacer(1, 0.3 * inch))
				summary_style = ParagraphStyle(
					'SummaryTitle',
					parent=styles['Heading2'],
					fontSize=14,
					spaceAfter=10
				)
				elements.append(Paragraph('Summary Statistics', summary_style))
				
				summary_data = [
					['Metric', 'Count', 'Percentage'],
					['Present', str(summary['present']), f"{summary['present_pct']:.1f}%"],
					['Absent', str(summary['absent']), f"{summary['absent_pct']:.1f}%"],
					['Late', str(summary['late']), '-'],
					['Excused', str(summary['excused']), '-'],
					['Total', str(summary['total']), '100%'],
				]
				
				summary_table = Table(summary_data)
				summary_table.setStyle(TableStyle([
					('BACKGROUND', (0, 0), (-1, 0), colors.grey),
					('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
					('ALIGN', (0, 0), (-1, -1), 'CENTER'),
					('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
					('FONTSIZE', (0, 0), (-1, 0), 10),
					('BOTTOMPADDING', (0, 0), (-1, 0), 12),
					('GRID', (0, 0), (-1, -1), 1, colors.black),
				]))
				elements.append(summary_table)
		
		# Build PDF
		doc.build(elements)
		pdf_data = buffer.getvalue()
		buffer.close()
		
		response = HttpResponse(content_type='application/pdf')
		response['Content-Disposition'] = f'attachment; filename="attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
		response.write(pdf_data)
		
		return response
